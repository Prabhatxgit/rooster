import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Page Configuration
st.set_page_config(
    page_title="Inbound Roster Automator",
    page_icon="📅",
    layout="wide"
)

# Custom CSS for better UI
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stDownloadButton button {
        background-color: #4CAF50;
        color: white;
        font-weight: bold;
        padding: 0.5rem 2rem;
    }
    </style>
""", unsafe_allow_html=True)

# Title
st.markdown('<div class="main-header">📅 Inbound Roster Automator</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Upload your employee data and generate clean rosters in seconds</div>', unsafe_allow_html=True)

# Sidebar Configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Upload Employee Data File",
        type=['csv', 'xlsx'],
        help="Upload your CSV or Excel file containing employee information"
    )
    
    st.divider()
    
    # Date input for start date
    default_date = datetime.now().replace(day=1) + timedelta(days=32)
    default_date = default_date.replace(day=1)
    
    start_date = st.date_input(
        "Select Roster Start Date",
        value=default_date,
        help="Choose the first day of the roster period"
    )
    
    st.divider()
    
    # Shift Logic Information
    with st.expander("ℹ️ Shift Logic Rules"):
        st.markdown("""
        **Week Offs (WO):**
        - Even rows: Saturday & Sunday
        - Odd rows: Tuesday & Wednesday
        
        **Shift Assignment:**
        - Even dates: 'Day' shift
        - Odd dates: 'Night' shift
        """)


def clean_employee_data(df):
    """
    Clean the input dataframe by removing garbage columns and extracting necessary fields.
    """
    # If it's an Excel file with header in first row, handle it
    if 'Employee ID' in df.iloc[0].values or 'Employee ID' in str(df.iloc[0].values):
        # First row contains the actual headers
        new_columns = df.iloc[0].tolist()
        df = df[1:].reset_index(drop=True)
        df.columns = new_columns
    
    # Remove columns that are unnamed or have "2" as name (garbage columns)
    columns_to_keep = []
    for col in df.columns:
        col_str = str(col).strip()
        # Keep only meaningful columns
        if col_str not in ['2', 'nan', 'Unnamed'] and not col_str.startswith('Unnamed:'):
            if col_str in ['Employee ID', 'NAME', 'Department', 'User ID', 'Status', 'WINS']:
                columns_to_keep.append(col)
    
    # If we don't have the expected columns, try to find them
    if not columns_to_keep:
        # Assume first few columns are Employee ID, User ID, NAME, Status, Department, WINS
        if len(df.columns) >= 3:
            df.columns = ['Employee ID', 'User ID', 'NAME', 'Status', 'Department', 'WINS'] + list(df.columns[6:])
            columns_to_keep = ['Employee ID', 'NAME', 'Department']
    
    # Keep only the necessary columns
    if columns_to_keep:
        df_clean = df[columns_to_keep].copy()
    else:
        # Fallback: assume first 3 columns are what we need
        df_clean = df.iloc[:, :3].copy()
        df_clean.columns = ['Employee ID', 'NAME', 'Department']
    
    # Ensure we have the required columns
    required_cols = ['Employee ID', 'NAME']
    for col in required_cols:
        if col not in df_clean.columns:
            st.error(f"Missing required column: {col}")
            return None
    
    # Add Department if missing
    if 'Department' not in df_clean.columns:
        df_clean['Department'] = 'Inbound'
    
    # Remove any empty rows
    df_clean = df_clean.dropna(subset=['Employee ID', 'NAME'], how='all')
    
    # Reset index
    df_clean = df_clean.reset_index(drop=True)
    
    return df_clean


def generate_roster(employee_df, start_date, num_days=30):
    """
    Generate roster with shift assignments based on the specified logic.
    """
    # Create a copy of employee data
    roster_df = employee_df.copy()
    
    # Generate date range
    date_range = [start_date + timedelta(days=i) for i in range(num_days)]
    
    # For each date, create a column and assign shifts
    for date in date_range:
        date_str = date.strftime('%Y-%m-%d')
        day_name = date.strftime('%A')
        day_of_month = date.day
        
        # Create column for this date
        column_name = f"{day_name[:3]} {date.strftime('%d-%b')}"
        
        # Initialize the column
        roster_df[column_name] = ''
        
        # Assign shifts based on row index
        for idx in roster_df.index:
            # Determine if this is an even or odd row (0-indexed)
            is_even_row = idx % 2 == 0
            
            # Week Off logic
            if is_even_row:
                # Even rows: Saturday & Sunday are WO
                if day_name in ['Saturday', 'Sunday']:
                    roster_df.at[idx, column_name] = 'WO'
                else:
                    # Assign shift based on date
                    roster_df.at[idx, column_name] = 'Day' if day_of_month % 2 == 0 else 'Night'
            else:
                # Odd rows: Tuesday & Wednesday are WO
                if day_name in ['Tuesday', 'Wednesday']:
                    roster_df.at[idx, column_name] = 'WO'
                else:
                    # Assign shift based on date
                    roster_df.at[idx, column_name] = 'Day' if day_of_month % 2 == 0 else 'Night'
    
    return roster_df


def style_excel(df, output_path):
    """
    Apply styling to the Excel file for better readability.
    """
    # Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, sheet_name='Roster', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Roster']
    
    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    day_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    night_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    wo_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Style data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Apply conditional formatting based on cell value
            if cell.value == 'Day':
                cell.fill = day_fill
            elif cell.value == 'Night':
                cell.fill = night_fill
            elif cell.value == 'WO':
                cell.fill = wo_fill
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    writer.close()


# Main Application Logic
if uploaded_file is not None:
    try:
        # Read the uploaded file
        if uploaded_file.name.endswith('.csv'):
            df_raw = pd.read_csv(uploaded_file)
        else:
            df_raw = pd.read_excel(uploaded_file)
        
        st.success(f"✅ File uploaded successfully! Found {len(df_raw)} rows.")
        
        # Clean the data
        with st.spinner("🧹 Cleaning employee data..."):
            df_clean = clean_employee_data(df_raw)
        
        if df_clean is not None and len(df_clean) > 0:
            # Display cleaned data
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("👥 Cleaned Employee Data")
                st.dataframe(
                    df_clean,
                    use_container_width=True,
                    height=300
                )
                st.caption(f"Total Employees: {len(df_clean)}")
            
            # Generate roster
            with st.spinner("⚙️ Generating roster..."):
                roster_df = generate_roster(df_clean, start_date, num_days=30)
            
            with col2:
                st.subheader("📅 Generated Roster Preview")
                # Show only first few columns for preview
                preview_cols = min(8, len(roster_df.columns))
                st.dataframe(
                    roster_df.iloc[:, :preview_cols],
                    use_container_width=True,
                    height=300
                )
                st.caption(f"Showing {preview_cols} of {len(roster_df.columns)} columns")
            
            # Full roster view
            st.divider()
            st.subheader("📊 Complete Roster")
            st.dataframe(
                roster_df,
                use_container_width=True,
                height=400
            )
            
            # Statistics
            st.divider()
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Employees", len(roster_df))
            with col2:
                st.metric("Roster Days", 30)
            with col3:
                start_str = start_date.strftime('%b %d, %Y')
                st.metric("Start Date", start_str)
            with col4:
                end_date = start_date + timedelta(days=29)
                end_str = end_date.strftime('%b %d, %Y')
                st.metric("End Date", end_str)
            
            # Download section
            st.divider()
            st.subheader("💾 Download Your Roster")
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                # Create Excel file in memory
                output = BytesIO()
                
                # Create styled Excel
                temp_path = '/tmp/roster_output.xlsx'
                style_excel(roster_df, temp_path)
                
                # Read the file back
                with open(temp_path, 'rb') as f:
                    excel_data = f.read()
                
                # Download button
                filename = f"Roster_{start_date.strftime('%B_%Y')}.xlsx"
                st.download_button(
                    label="⬇️ Download Excel Roster",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                st.info("📌 The Excel file includes:\n- Color-coded shifts\n- Formatted headers\n- Auto-sized columns")
        
        else:
            st.error("❌ Could not process the uploaded file. Please check the file format.")
    
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.exception(e)

else:
    # Welcome message when no file is uploaded
    st.info("👆 Please upload an employee data file from the sidebar to get started.")
    
    # Instructions
    with st.expander("📖 How to Use This Tool"):
        st.markdown("""
        ### Step-by-Step Guide
        
        1. **Upload Your File**: Click the file uploader in the sidebar and select your CSV or Excel file
        2. **Select Start Date**: Choose the first day of your roster period
        3. **Review Data**: Check the cleaned employee data and generated roster preview
        4. **Download**: Click the download button to get your formatted Excel roster
        
        ### Input File Requirements
        
        Your file should contain these columns:
        - `Employee ID`: Unique identifier for each employee
        - `NAME`: Employee name
        - `Department`: Department name (optional)
        
        The tool will automatically:
        - Remove garbage columns (like repeated "2" columns)
        - Clean empty columns
        - Extract only necessary information
        
        ### Shift Assignment Rules
        
        **Week Offs:**
        - Employees in even rows (0, 2, 4...): Saturday & Sunday
        - Employees in odd rows (1, 3, 5...): Tuesday & Wednesday
        
        **Shifts:**
        - Even dates (2nd, 4th, 6th...): Day shift
        - Odd dates (1st, 3rd, 5th...): Night shift
        """)
    
    # Sample data showcase
    with st.expander("👀 See Example Output"):
        st.markdown("### Sample Roster Output")
        st.image("https://via.placeholder.com/800x300/4472C4/FFFFFF?text=Color-Coded+Roster+with+Day,+Night,+and+WO+Shifts", 
                 caption="Your roster will be color-coded for easy reading")

# Footer
st.divider()
st.markdown(
    "<div style='text-align: center; color: #666; padding: 1rem;'>"
    "Built with ❤️ using Streamlit | Roster Automation Tool v1.0"
    "</div>",
    unsafe_allow_html=True
)
